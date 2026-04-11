"""
report/excel.py — Generate Excel (.xlsx) property financial reports.

Sheet layout "Finální sestava":
  Rows 1-4 : Config header (kurz, provize, balíčky, property title)
  Row  5   : Column headers
  Row  6+  : Reservation rows (full calculations) + payment sub-rows (grey)
             ─── PLATBY PŘENESENÉ Z PŘEDCHOZÍCH MĚSÍCŮ (separator, if any) ───
             Transferred payment rows from previous months (blue)
  Last rows: Totals row → Summary block (accrual + bank reconciliation)

Row types written by _build_final_sheet:
  RESERVATION  — white/coloured by verification status, full calculations
  PAYMENT      — grey sub-row under each reservation: REF + CZK received + date
  SEPARATOR    — full-width dark divider before transferred section
  TRANSFERRED  — blue rows: payment received after cutoff from a previous month
"""

import os
import re
from calendar import month_name as _MONTH_NAMES
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from report.summary import build_report_summary
from report.verifier import (
    STATUS_MATCHED, STATUS_ROZDIL, STATUS_CHYBI_CSV,
    STATUS_CHYBI_HOSTIFY, STATUS_ZRUSENO, STATUS_KE_KONTROLE,
)

# --------------------------------------------------------------------------- #
#  Colours                                                                     #
# --------------------------------------------------------------------------- #

_FILL_MATCHED    = PatternFill("solid", fgColor="C6EFCE")   # green
_FILL_ROZDIL     = PatternFill("solid", fgColor="FFEB9C")   # yellow
_FILL_CHYBI      = PatternFill("solid", fgColor="FFC7CE")   # red
_FILL_ZRUSENO    = PatternFill("solid", fgColor="D9D9D9")   # grey
_FILL_HEADER     = PatternFill("solid", fgColor="DAEEF3")   # light blue
_FILL_TOTALS     = PatternFill("solid", fgColor="EBF1DE")   # light green
_FILL_PAYMENT    = PatternFill("solid", fgColor="F2F2F2")   # light grey — payment sub-row
_FILL_TRANSFERRED= PatternFill("solid", fgColor="DCE6F1")   # blue — transferred row
_FILL_SEPARATOR  = PatternFill("solid", fgColor="595959")   # dark grey — separator
_FILL_BANK_OK    = PatternFill("solid", fgColor="C6EFCE")   # green — DORAZILO
_FILL_BANK_MISS  = PatternFill("solid", fgColor="FFC7CE")   # red — CHYBÍ
_FILL_SUMMARY    = PatternFill("solid", fgColor="F2F2F2")

_FILL_KE_KONTROLE = PatternFill("solid", fgColor="FFEB9C")   # yellow — needs review

STATUS_FILLS = {
    STATUS_MATCHED:       _FILL_MATCHED,
    STATUS_ROZDIL:        _FILL_ROZDIL,
    STATUS_CHYBI_CSV:     _FILL_CHYBI,
    STATUS_CHYBI_HOSTIFY: _FILL_CHYBI,
    STATUS_ZRUSENO:       _FILL_ZRUSENO,
    STATUS_KE_KONTROLE:   _FILL_KE_KONTROLE,
}

# --------------------------------------------------------------------------- #
#  Number formats                                                              #
# --------------------------------------------------------------------------- #

FMT_CZK  = '#,##0.00\\ "Kč"'
FMT_EUR  = '#,##0.00\\ "€"'
FMT_NUM  = '#,##0.00'
FMT_INT  = '#,##0'
FMT_PCT  = '0.00%'
FMT_TEXT = '@'

# --------------------------------------------------------------------------- #
#  Fonts & borders                                                             #
# --------------------------------------------------------------------------- #

_FONT_HEADER     = Font(bold=True, size=9)
_FONT_TITLE      = Font(bold=True, size=11)
_FONT_NORMAL     = Font(size=9)
_FONT_SMALL      = Font(size=8, italic=True)
_FONT_TOTALS     = Font(bold=True, size=9)
_FONT_SEPARATOR  = Font(bold=True, size=9, color="FFFFFF")
_FONT_TRANSFERRED= Font(size=9, color="17375E")

_THIN  = Side(style="thin",   color="BFBFBF")
_THICK = Side(style="medium", color="595959")
_BORDER_CELL   = Border(left=_THIN, right=_THIN, top=_THIN,  bottom=_THIN)
_BORDER_HEADER = Border(left=_THIN, right=_THIN, top=_THICK, bottom=_THICK)
_BORDER_NONE   = Border()


def _align(horizontal="center", vertical="center", wrap=False):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


# --------------------------------------------------------------------------- #
#  Column definitions                                                          #
#                                                                             #
#  Columns A–Q  : Calculation columns (reservation rows)                      #
#  Columns R–S  : Verification status + comment (reservation rows)            #
#  Columns T–W  : Bank payment info (payment sub-rows + transferred rows)     #
# --------------------------------------------------------------------------- #

# (header_text, data_key, width, number_format, align)
_FINAL_COLS = [
    # ── Identification ──────────────────────────────────────────────────────
    ("Pořadí",              "order",                    5,  FMT_INT,  "center"),  # A
    ("Jméno hosta",         "guest_name",               24, FMT_TEXT, "left"),    # B
    ("Pobyt",               "stay_label",               13, FMT_TEXT, "center"),  # C
    ("Dní",                 "nights",                   5,  FMT_INT,  "center"),  # D
    ("Osob P",              "adults",                   7,  FMT_INT,  "center"),  # E
    ("Osob O",              "children_infants",         7,  FMT_INT,  "center"),  # F
    # ── Calculations ────────────────────────────────────────────────────────
    ("Místní popl.",        "city_tax_czk",             13, FMT_CZK,  "right"),   # G
    ("Provize",             "provize_czk",              13, FMT_CZK,  "right"),   # H
    ("DPH provize",         "dph_provize_czk",          13, FMT_CZK,  "right"),   # I
    ("Výplata EUR",         "payout_eur",               13, FMT_EUR,  "right"),   # J
    ("Vyplaceno CZK",       "payout_czk",               14, FMT_CZK,  "right"),   # K
    ("Úklid EUR",           "cleaning_fee_eur",         11, FMT_EUR,  "right"),   # L
    ("Úklid CZK",           "uklid_czk",                13, FMT_CZK,  "right"),   # M
    ("Balíčky",             "balicky_czk",              11, FMT_CZK,  "right"),   # N
    ("DPH úklid+bal.",      "dph_uklid_balicky_czk",   15, FMT_CZK,  "right"),   # O
    ("Příprava pokoje",     "priprava_pokoje_czk",      16, FMT_CZK,  "right"),   # P
    ("Cena ubytování",      "cena_ubytovani_czk",       16, FMT_CZK,  "right"),   # Q
    # ── Verification ────────────────────────────────────────────────────────
    ("Status",              "verification_status",      11, FMT_TEXT, "center"),  # R
    ("Komentář",            "comment",                  28, FMT_TEXT, "left"),    # S
    # ── Bank payment (sub-rows / transferred) ────────────────────────────────
    ("Platební ref",        "payout_gref",              24, FMT_TEXT, "left"),    # T
    ("Přijato CZK",         "bank_amount_czk",          14, FMT_CZK,  "right"),   # U
    ("Banka datum",         "bank_datum",               12, FMT_TEXT, "center"),  # V
    ("Platba",              "bank_status",              11, FMT_TEXT, "center"),  # W
]

_N_COLS = len(_FINAL_COLS)  # 23

# Columns whose values are summed in the totals row (reservation rows only)
_SUM_KEYS = {
    "city_tax_czk", "provize_czk", "dph_provize_czk",
    "payout_eur", "payout_czk", "cleaning_fee_eur",
    "uklid_czk", "balicky_czk", "dph_uklid_balicky_czk",
    "priprava_pokoje_czk", "cena_ubytovani_czk",
    "nights", "adults",
}

# Column index lookup (1-based)
_COL_IDX = {key: i + 1 for i, (_, key, _, _, _) in enumerate(_FINAL_COLS)}


# --------------------------------------------------------------------------- #
#  Low-level cell helpers                                                      #
# --------------------------------------------------------------------------- #

def _set_col_widths(ws):
    for i, (_, _, width, _, _) in enumerate(_FINAL_COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _c(ws, row, col, value=None, *, font=None, fill=None, fmt=None,
       align=None, border=_BORDER_CELL):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if fmt:    cell.number_format = fmt
    if align:  cell.alignment = align
    if border is not None: cell.border = border
    return cell


def _fill_row(ws, row_num: int, fill, n_cols: int = _N_COLS):
    """Apply fill + border to an entire row."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.border = _BORDER_CELL


# --------------------------------------------------------------------------- #
#  Row writers                                                                 #
# --------------------------------------------------------------------------- #

def _write_header_row(ws, row_num: int):
    for col_idx, (header, _, _, _, halign) in enumerate(_FINAL_COLS, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.border = _BORDER_HEADER
        cell.alignment = _align(horizontal=halign, wrap=True)


def _write_reservation_row(ws, row_num: int, row_data: dict):
    """Write one reservation row (cols A–S filled, T–W empty)."""
    status = row_data.get("verification_status", "")
    row_fill = STATUS_FILLS.get(status)

    for col_idx, (_, key, _, fmt, halign) in enumerate(_FINAL_COLS, start=1):
        # Bank columns (T–W) are written by the payment sub-row, not here
        if key in ("payout_gref", "bank_amount_czk", "bank_datum", "bank_status"):
            val = None
        else:
            val = row_data.get(key)

        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.number_format = fmt
        cell.font = _FONT_NORMAL
        cell.border = _BORDER_CELL
        cell.alignment = _align(horizontal=halign)

        # Status cell gets its own per-cell fill
        if key == "verification_status" and row_fill:
            cell.fill = row_fill
        elif row_fill:
            cell.fill = row_fill


def _write_payment_subrow(ws, row_num: int, row_data: dict):
    """
    Write a payment sub-row (grey):
      A = "  ↳"
      B = REF string (gref or NO-ref)
      C–S = empty
      T = payout_gref (same as B but for alignment)
      U = bank_amount_czk
      V = bank_datum
      W = bank_status symbol
    """
    gref        = row_data.get("payout_gref", "") or ""
    amount_czk  = row_data.get("bank_amount_czk")
    bank_datum  = row_data.get("bank_datum", "")
    bank_status = row_data.get("bank_status", "")

    status_symbol = "✓ DORAZILO" if bank_status == "DORAZILO" else "⚠ ČEKÁ/PŘENOS"
    status_fill   = _FILL_BANK_OK if bank_status == "DORAZILO" else _FILL_BANK_MISS

    # Fill entire row light grey first
    _fill_row(ws, row_num, _FILL_PAYMENT)

    _c(ws, row_num, 1,  "  ↳",        font=_FONT_SMALL, fill=_FILL_PAYMENT)
    _c(ws, row_num, 2,  gref,          font=_FONT_SMALL, fill=_FILL_PAYMENT,
       align=_align("left"))
    # Cols 3-19 stay grey (already filled)
    for col in range(3, 20):
        ws.cell(row=row_num, column=col).fill = _FILL_PAYMENT
        ws.cell(row=row_num, column=col).border = _BORDER_CELL

    # T: repeat REF (cleaner alignment in bank section)
    _c(ws, row_num, _COL_IDX["payout_gref"], gref,
       font=_FONT_SMALL, fill=_FILL_PAYMENT, align=_align("left"))

    # U: amount
    if amount_czk is not None:
        _c(ws, row_num, _COL_IDX["bank_amount_czk"], amount_czk,
           font=_FONT_SMALL, fill=_FILL_PAYMENT, fmt=FMT_CZK,
           align=_align("right"))
    else:
        ws.cell(row=row_num, column=_COL_IDX["bank_amount_czk"]).fill = _FILL_PAYMENT

    # V: date
    _c(ws, row_num, _COL_IDX["bank_datum"], bank_datum or "—",
       font=_FONT_SMALL, fill=_FILL_PAYMENT, align=_align("center"))

    # W: status
    _c(ws, row_num, _COL_IDX["bank_status"], status_symbol,
       font=_FONT_SMALL, fill=status_fill, align=_align("center"))


def _write_separator_row(ws, row_num: int, label: str):
    """Full-width dark separator row."""
    ws.merge_cells(
        start_row=row_num, start_column=1,
        end_row=row_num, end_column=_N_COLS,
    )
    cell = ws.cell(row=row_num, column=1, value=label)
    cell.font = _FONT_SEPARATOR
    cell.fill = _FILL_SEPARATOR
    cell.alignment = _align("left")
    cell.border = _BORDER_CELL
    ws.row_dimensions[row_num].height = 14


def _write_transferred_row(ws, row_num: int, p: dict):
    """
    Blue row for a payment received from a previous month.
      A = "↩"
      B = "z [month YYYY]"
      C = guest_name
      D = stay_label (original dates)
      E–S = empty
      T = gref
      U = bank_amount_czk
      V = bank_datum
      W = "✓ PŘIJATO"
    """
    _fill_row(ws, row_num, _FILL_TRANSFERRED)

    orig_month = p.get("original_month", 0)
    orig_year  = p.get("original_year", 0)
    try:
        month_label = f"z {_MONTH_NAMES[orig_month]} {orig_year}"
    except (IndexError, KeyError):
        month_label = f"z {orig_month:02d}/{orig_year}"

    _c(ws, row_num, 1, "↩",          font=_FONT_TRANSFERRED, fill=_FILL_TRANSFERRED)
    _c(ws, row_num, 2, month_label,   font=_FONT_TRANSFERRED, fill=_FILL_TRANSFERRED,
       align=_align("left"))
    _c(ws, row_num, 3, p.get("guest_name", ""),
       font=_FONT_TRANSFERRED, fill=_FILL_TRANSFERRED, align=_align("left"))
    _c(ws, row_num, 4, p.get("stay_label", ""),
       font=_FONT_TRANSFERRED, fill=_FILL_TRANSFERRED, align=_align("center"))
    for col in range(5, 20):
        ws.cell(row=row_num, column=col).fill = _FILL_TRANSFERRED
        ws.cell(row=row_num, column=col).border = _BORDER_CELL

    _c(ws, row_num, _COL_IDX["payout_gref"], p.get("gref", ""),
       font=_FONT_TRANSFERRED, fill=_FILL_TRANSFERRED, align=_align("left"))

    amount = p.get("bank_amount_czk")
    if amount is not None:
        _c(ws, row_num, _COL_IDX["bank_amount_czk"], amount,
           font=Font(bold=True, size=9, color="17375E"),
           fill=_FILL_TRANSFERRED, fmt=FMT_CZK, align=_align("right"))

    _c(ws, row_num, _COL_IDX["bank_datum"], p.get("bank_datum", ""),
       font=_FONT_TRANSFERRED, fill=_FILL_TRANSFERRED, align=_align("center"))

    _c(ws, row_num, _COL_IDX["bank_status"], "✓ PŘIJATO",
       font=Font(bold=True, size=9, color="17375E"),
       fill=_FILL_BANK_OK, align=_align("center"))


def _write_totals_row(ws, row_num: int, res_rows: list[dict]):
    """Sum row — only over reservation rows (not sub-rows or transferred)."""
    ws.cell(row=row_num, column=1, value="CELKEM").font = _FONT_TOTALS
    for col in range(1, _N_COLS + 1):
        ws.cell(row=row_num, column=col).fill  = _FILL_TOTALS
        ws.cell(row=row_num, column=col).border = _BORDER_CELL

    ws.cell(row=row_num, column=1).alignment = _align("center")

    for col_idx, (_, key, _, fmt, halign) in enumerate(_FINAL_COLS, start=2):
        if key in _SUM_KEYS:
            total = sum(float(r[key]) for r in res_rows if r.get(key) is not None)
            cell = ws.cell(row=row_num, column=col_idx, value=round(total, 2))
            cell.number_format = fmt
            cell.font = _FONT_TOTALS
            cell.alignment = _align(horizontal=halign)


# --------------------------------------------------------------------------- #
#  Summary block                                                               #
# --------------------------------------------------------------------------- #

def _write_summary_block(
    ws,
    start_row: int,
    summary: dict,
    year: int,
    month: int,
):
    """
    Write the financial summary block below the totals row.
    Two sections:
      1. Accrual (výnosy): standard Rentero / klient summary
      2. Bank reconciliation: potvrzeno / čeká / přijato z min. měsíců
    """
    r = start_row
    def _lbl(row_r, label, value, fmt=FMT_CZK, bold=False):
        ws.cell(row=row_r, column=2, value=label).font = (
            Font(bold=True, size=9) if bold else _FONT_NORMAL
        )
        cell = ws.cell(row=row_r, column=4, value=value)
        cell.number_format = fmt
        cell.font = Font(bold=bold, size=9)
        cell.alignment = _align("right")
        ws.cell(row=row_r, column=2).fill = _FILL_SUMMARY
        ws.cell(row=row_r, column=3).fill = _FILL_SUMMARY
        ws.cell(row=row_r, column=4).fill = _FILL_SUMMARY
        ws.cell(row=row_r, column=5).fill = _FILL_SUMMARY

    # ── Section 1: Rentero / klient ──────────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(row=r, column=2,
            value=f"Summary {month:02d}/{year} — Rentero").font = _FONT_TITLE
    r += 1
    _lbl(r, "Příjem Rentero (management fee)",            summary["rentero_fee_czk"]); r += 1
    _lbl(r, "DPH z příjmu Rentero",                       summary["vat_rentero_fee_czk"]); r += 1
    _lbl(r, "Příjem Rentero — příprava pokoje (vč. DPH)", summary["rentero_room_prep_with_vat_czk"]); r += 1
    r += 1

    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(row=r, column=2,
            value=f"Klient Summary {month:02d}/{year}").font = _FONT_TITLE
    r += 1
    _lbl(r, "Vyplaceno platformami celkem",               summary["gross_payout_czk"]); r += 1
    _lbl(r, "Příjem hrubý klient (Cena ubytování)",       summary["client_gross_income_czk"]); r += 1
    _lbl(r, "Klientovi bude vyplaceno před výdaji",       summary["client_payout_before_expenses_czk"]); r += 1
    if summary["expenses_total_czk"]:
        _lbl(r, "Výdaje",                                 summary["expenses_total_czk"]); r += 1
    _lbl(r, "Klientovi bude vyplaceno po výdajích",       summary["client_payout_after_expenses_czk"], bold=True); r += 1
    _lbl(r, "DPH přefakturace klient",                    summary["dph_prefakturace_klient_czk"]); r += 1
    r += 1

    # ── Section 2: Bank reconciliation ──────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(row=r, column=2,
            value="Bankovní reconciliace").font = _FONT_TITLE
    r += 1
    _lbl(r, "Potvrzeno bankou (DORAZILO)",                   summary["bank_confirmed_czk"], bold=True); r += 1
    _lbl(r, "Čeká na platbu → přenese se do příštího měsíce", summary["bank_pending_czk"]); r += 1
    if summary["bank_transferred_czk"]:
        _lbl(r, "Přijato z předchozích měsíců",             summary["bank_transferred_czk"], bold=True); r += 1
    _lbl(r, "Celkem přijato tento měsíc (potvrzeno + přeneseno)",
         summary["bank_received_this_month_czk"], bold=True);                         r += 1


# --------------------------------------------------------------------------- #
#  Main sheet builder                                                          #
# --------------------------------------------------------------------------- #

def _build_final_sheet(
    wb: openpyxl.Workbook,
    rows: list[dict],
    totals: dict,
    property_config: dict,
    year: int,
    month: int,
    transferred_rows: list[dict] | None = None,
    expenses: list[dict] | None = None,
) -> None:
    if transferred_rows is None:
        transferred_rows = []
    if expenses is None:
        expenses = []

    ws = wb.active
    ws.title = "Finální sestava"

    avg_kurz = _avg_kurz(rows)

    # ── Rows 1-4: config header ──────────────────────────────────────────────
    ws.cell(row=1, column=2, value="kurz eura (průměr)").font = _FONT_HEADER
    ws.cell(row=1, column=3, value=avg_kurz).number_format = FMT_NUM
    ws.cell(row=1, column=4, value="Rentero provize").font = _FONT_HEADER
    ws.cell(row=1, column=5,
            value=property_config.get("rentero_commission", 0.15)).number_format = FMT_PCT

    ws.cell(row=2, column=2, value="Balíček hosta").font = _FONT_HEADER
    ws.cell(row=2, column=3, value="Premium/osoba").font = Font(size=9, italic=True)
    ws.cell(row=2, column=4,
            value=property_config.get("balicky_per_person", 0)).number_format = FMT_CZK

    title = f"{property_config.get('display_name', '')} — {month:02d}/{year}"
    ws.cell(row=3, column=1, value=title).font = _FONT_TITLE

    ws.cell(row=4, column=2,
            value="Údaje nutné k fakturaci").font = Font(bold=True, size=9, italic=True)
    ws.cell(row=4, column=20,
            value="Bankovní platby").font = Font(bold=True, size=9, italic=True)

    # ── Row 5: column headers ─────────────────────────────────────────────────
    _write_header_row(ws, 5)

    # ── Rows 6+: reservations + payment sub-rows ──────────────────────────────
    current_row = 6
    res_rows = []  # reservation rows only (for totals)

    for row_data in rows:
        # Skip CHYBÍ_V_HOSTIFY rows with no meaningful data
        if row_data.get("verification_status") == STATUS_CHYBI_HOSTIFY:
            if not row_data.get("guest_name") and not row_data.get("check_in"):
                continue

        _write_reservation_row(ws, current_row, row_data)
        res_rows.append(row_data)
        current_row += 1

        # Payment sub-row (only for Airbnb / Booking — not N/A)
        if row_data.get("bank_status") not in (None, "N/A"):
            _write_payment_subrow(ws, current_row, row_data)
            current_row += 1

    # ── Transferred payments from previous months ─────────────────────────────
    if transferred_rows:
        current_row += 1  # blank spacer
        _write_separator_row(
            ws, current_row,
            "  PLATBY PŘENESENÉ Z PŘEDCHOZÍCH MĚSÍCŮ"
        )
        current_row += 1

        for p in transferred_rows:
            _write_transferred_row(ws, current_row, p)
            current_row += 1

    # ── Totals row ────────────────────────────────────────────────────────────
    current_row += 1
    _write_totals_row(ws, current_row, res_rows)
    totals_row = current_row

    # ── Summary block ─────────────────────────────────────────────────────────
    summary = build_report_summary(
        res_rows,
        property_config,
        expenses=expenses,
        transferred_rows=transferred_rows,
    )
    _write_summary_block(ws, totals_row + 2, summary, year, month)

    # ── Finalize ──────────────────────────────────────────────────────────────
    _set_col_widths(ws)
    ws.freeze_panes = "B6"
    ws.row_dimensions[5].height = 28   # header row taller


# --------------------------------------------------------------------------- #
#  Public API                                                                  #
# --------------------------------------------------------------------------- #

def _avg_kurz(rows: list[dict]) -> float:
    rates = [r["kurz"] for r in rows if r.get("kurz")]
    return round(sum(rates) / len(rates), 3) if rates else 0.0


def _slug_to_filename(slug: str) -> str:
    return re.sub(r"[^A-Za-z0-9_]", "_", slug)


def _check_file_locked(path: str) -> bool:
    if not os.path.exists(path):
        return False
    try:
        os.rename(path, path)
        return False
    except OSError:
        return True


def write_property_report(
    rows: list[dict],
    totals: dict,
    property_config: dict,
    year: int,
    month: int,
    output_dir: str,
    *,
    overwrite: bool = False,
    transferred_rows: list[dict] | None = None,
    expenses: list[dict] | None = None,
) -> str:
    """
    Generate one .xlsx file for a property-month.

    Args:
        rows: CalculatedRow dicts (all rows)
        totals: from calculator.calculate_totals_with_config()
        property_config: single property config dict with 'slug'
        year, month: target period
        output_dir: directory to write the file into
        overwrite: if False and file exists, raises FileExistsError
        transferred_rows: resolved pending payments from previous months

    Returns:
        Absolute path to the generated file.
    """
    os.makedirs(output_dir, exist_ok=True)
    slug = property_config.get("slug", "unknown")
    filename = f"Rentero_pracovni_{month:02d}{year}_{_slug_to_filename(slug)}.xlsx"
    output_path = os.path.abspath(os.path.join(output_dir, filename))

    if os.path.exists(output_path) and not overwrite:
        raise FileExistsError(
            f"Output file already exists: {output_path}. Use --overwrite to replace."
        )
    if _check_file_locked(output_path):
        raise OSError(
            f"File is open in another application: {output_path}. Close it and retry."
        )

    wb = openpyxl.Workbook()
    _build_final_sheet(wb, rows, totals, property_config, year, month,
                       transferred_rows=transferred_rows or [],
                       expenses=expenses or [])
    wb.save(output_path)
    return output_path
